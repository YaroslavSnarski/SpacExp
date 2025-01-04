import openpyxl
from .base_processor import FileProcessor
import os
import time
import xlrd
from .logging_config import setup_logging

# логгирование
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
process_logger, error_logger = setup_logging(project_root)

class ExcelProcessor(FileProcessor):
    """
    Класс для обработки Excel файлов формата .xlsx.

    Наследуется:
        FileProcessor: Базовый класс для обработки файлов.

    Методы:
        process(filepath): Извлекает общее количество листов в файле Excel.
    """
    def process(self, filepath):
        """
        Обрабатывает Excel файл (.xlsx) и извлекает количество листов.

        Args:
            filepath (str): Путь к Excel файлу.

        Returns:
            dict: Словарь с информацией о файле, включая:
                - num_sheets (int): Количество листов в книге.

            В случае ошибки возвращает словарь с ключом "error" и описанием ошибки.
        """
        try:
            start_time = time.time()
            file_info = self.get_generic_info(filepath)

            workbook = openpyxl.load_workbook(filepath)
            file_info.update({"num_sheets": len(workbook.sheetnames)})

            elapsed_time = time.time() - start_time
            process_logger.info(f"Excel processed: {filepath} in {elapsed_time:.2f} seconds")

            return file_info
        except Exception as e:
            error_logger.error(f"Error processing Excel {filepath}: {e}")
            return {"error": str(e)}

class ExcelProcessorWeb(FileProcessor):
    """
    Класс для обработки Excel файлов форматов .xls и .xlsx.

    Наследуется:
        FileProcessor: Базовый класс для обработки файлов.

    Методы:
        process(filepath): Извлекает количество листов и их названия из файла Excel.
    """
    def process(self, filepath):
        """
        Обрабатывает Excel файл (.xls или .xlsx) и извлекает количество листов и их названия.

        Args:
            filepath (str): Путь к Excel файлу.

        Returns:
            dict: Словарь с информацией о файле, включая:
                - num_sheets (int): Количество листов в книге.
                - sheet_names (list): Список названий листов.

            В случае ошибки возвращает словарь с ключом "error" и описанием ошибки.
        """
        try:
            start_time = time.time()
            file_info = self.get_generic_info(filepath)
            extension = filepath.split('.')[-1].lower()

            if extension == "xls":
                # Обработка .xls файлов с помощью xlrd
                workbook = xlrd.open_workbook(filepath)
                sheet_names = workbook.sheet_names()
            elif extension == "xlsx":
                # Обработка .xlsx файлов с помощью openpyxl
                workbook = openpyxl.load_workbook(filepath)
                sheet_names = workbook.sheetnames
            else:
                raise ValueError("Unsupported Excel file format")

            file_info.update({
                "num_sheets": len(sheet_names),
                "sheet_names": sheet_names  # adding sheets names to results
            })

            elapsed_time = time.time() - start_time
            process_logger.info(f"Excel processed: {filepath} in {elapsed_time:.2f} seconds")

            return file_info
        except Exception as e:
            error_logger.error(f"Error processing Excel {filepath}: {e}")
            return {"error": str(e)}
